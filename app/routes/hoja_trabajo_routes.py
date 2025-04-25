# app/routes/hoja_trabajo_routes.py
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from app.models.hoja_trabajo import HojaTrabajo, CapituloHoja, PartidaHoja
from app.models.presupuesto import Presupuesto, Capitulo, Partida
from app.models.proyecto import Proyecto
from app.models.cliente import Cliente
from app.models.proveedor import Proveedor
from app.models.proveedor_partida import ProveedorPartida
from app.services.db_service import get_all, get_by_id, create, update, delete
from app.services.pdf_service import generar_pdf_hoja_trabajo
from app import db
from datetime import datetime
import re
import os
import tempfile

# Función para sanitizar texto HTML (elimina todas las etiquetas)
def sanitize_html(text):
    if text is None:
        return ''
    return re.sub(r'<[^>]*>', '', text)

# Función para limpiar HTML pero preservar etiquetas básicas de formato
def clean_html(text):
    if text is None:
        return ''
    # Definir etiquetas permitidas
    allowed_tags = ['p', 'br', 'strong', 'b', 'em', 'i', 'u', 's', 'ul', 'ol', 'li', 'blockquote']
    
    # Crear patrón regex para eliminar etiquetas no permitidas
    allowed_pattern = '|'.join(allowed_tags)
    pattern = r'<(?!/?({})[\s>]).*?>'.format(allowed_pattern)
    
    # Eliminar etiquetas que no estén en la lista de permitidas
    return re.sub(pattern, '', text)

hojas_trabajo_bp = Blueprint('hojas_trabajo', __name__, url_prefix='/hojas_trabajo')

@hojas_trabajo_bp.route('/')
def listar_hojas_trabajo():
    """Lista todas las hojas de trabajo junto con información del presupuesto y proyecto asociado"""
    hojas = db.session.query(
        HojaTrabajo, Presupuesto, Proyecto, Cliente
    ).join(
        Presupuesto, HojaTrabajo.id_presupuesto == Presupuesto.id
    ).join(
        Proyecto, Presupuesto.id_proyecto == Proyecto.id
    ).join(
        Cliente, Proyecto.id_cliente == Cliente.id
    ).order_by(
        HojaTrabajo.fecha.desc()
    ).all()
    
    return render_template('hojas_trabajo/lista.html', hojas=hojas)

@hojas_trabajo_bp.route('/por-presupuesto/<int:id_presupuesto>')
def hojas_por_presupuesto(id_presupuesto):
    """Lista todas las hojas de trabajo asociadas a un presupuesto específico"""
    presupuesto = get_by_id(Presupuesto, id_presupuesto)
    if not presupuesto:
        flash('Presupuesto no encontrado', 'danger')
        return redirect(url_for('presupuestos.listar_presupuestos'))
    
    proyecto = get_by_id(Proyecto, presupuesto.id_proyecto)
    cliente = get_by_id(Cliente, proyecto.id_cliente) if proyecto.id_cliente else None
    
    hojas = HojaTrabajo.query.filter_by(id_presupuesto=id_presupuesto).all()
    
    return render_template('hojas_trabajo/por_presupuesto.html', 
                          hojas=hojas, 
                          presupuesto=presupuesto,
                          proyecto=proyecto,
                          cliente=cliente)

@hojas_trabajo_bp.route('/nueva/<int:id_presupuesto>', methods=['GET', 'POST'])
def nueva_hoja(id_presupuesto):
    """Crea una nueva hoja de trabajo basada en un presupuesto específico"""
    presupuesto = get_by_id(Presupuesto, id_presupuesto)
    if not presupuesto:
        flash('Presupuesto no encontrado', 'danger')
        return redirect(url_for('presupuestos.listar_presupuestos'))
    
    proyecto = get_by_id(Proyecto, presupuesto.id_proyecto)
    cliente = get_by_id(Cliente, proyecto.id_cliente) if proyecto.id_cliente else None
    
    if request.method == 'POST':
        try:
            # Generar referencia automática basada en la referencia del presupuesto
            # Añadir "HT" al final de la referencia del presupuesto
            referencia = f"{presupuesto.referencia}HT"
            
            # Verificar si ya existe la referencia
            hoja_existente = HojaTrabajo.query.filter_by(referencia=referencia).first()
            if hoja_existente:
                flash(f'Ya existe una hoja de trabajo con la referencia {referencia}', 'danger')
                return redirect(url_for('hojas_trabajo.nueva_hoja', id_presupuesto=id_presupuesto))
            
            # Crear la hoja de trabajo
            data = {
                'id_presupuesto': id_presupuesto,
                'referencia': referencia,
                'fecha': datetime.utcnow(),
                'tipo_via': request.form.get('tipo_via') or presupuesto.tipo_via,
                'nombre_via': request.form.get('nombre_via') or presupuesto.nombre_via,
                'numero_via': request.form.get('numero_via') or presupuesto.numero_via,
                'puerta': request.form.get('puerta') or presupuesto.puerta,
                'codigo_postal': request.form.get('codigo_postal') or presupuesto.codigo_postal,
                'poblacion': request.form.get('poblacion') or presupuesto.poblacion,
                'titulo': sanitize_html(request.form.get('titulo')) or f"Hoja de trabajo para {presupuesto.titulo}",
                'notas': sanitize_html(request.form.get('notas', '')),
                'tecnico_encargado': request.form.get('tecnico_encargado') or presupuesto.tecnico_encargado or 'Sin asignar',
                'aprobacion': None,
                'fecha_aprobacion': None,
                'estado': 'Borrador'
            }
            
            # Crear la instancia de HojaTrabajo y guardarla
            hoja = HojaTrabajo(**data)
            db.session.add(hoja)
            db.session.flush()  # Para obtener el ID de la hoja
            
            # Copiar capítulos y partidas del presupuesto a la hoja de trabajo
            capitulos = Capitulo.query.filter_by(id_presupuesto=id_presupuesto).all()
            for cap in capitulos:
                nuevo_cap = CapituloHoja(
                    id_hoja=hoja.id,
                    numero=cap.numero,
                    descripcion=cap.descripcion
                )
                db.session.add(nuevo_cap)
            
            # Copiar partidas
            partidas = Partida.query.filter_by(id_presupuesto=id_presupuesto).all()
            for part in partidas:
                descripcion_limpia = sanitize_html(part.descripcion) if part.descripcion else ''
                
                # Usar clean_html en lugar de sanitize_html para preservar formato
                descripcion_con_formato = clean_html(part.descripcion) if part.descripcion else ''
                
                nueva_part = PartidaHoja(
                    id_hoja=hoja.id,
                    capitulo_numero=part.capitulo_numero,
                    descripcion=descripcion_con_formato,
                    unitario=part.unitario,
                    cantidad=part.cantidad,
                    precio=part.precio,
                    total=part.total,
                    margen=part.margen,
                    final=part.final,
                    id_proveedor_principal=None,
                    precio_proveedor=None
                )
                db.session.add(nueva_part)
            
            db.session.commit()
            flash('Hoja de trabajo creada correctamente', 'success')
            return redirect(url_for('hojas_trabajo.editar_hoja', id=hoja.id))
        
        except Exception as e:
            db.session.rollback()
            flash(f'Error al crear hoja de trabajo: {str(e)}', 'danger')
    
    # Renderizar el formulario para crear una nueva hoja de trabajo
    return render_template('hojas_trabajo/nueva.html', 
                          presupuesto=presupuesto,
                          proyecto=proyecto, 
                          cliente=cliente)

@hojas_trabajo_bp.route('/editar/<int:id>', methods=['GET', 'POST'])
def editar_hoja(id):
    """Edita una hoja de trabajo existente"""
    hoja = get_by_id(HojaTrabajo, id)
    if not hoja:
        flash('Hoja de trabajo no encontrada', 'danger')
        return redirect(url_for('hojas_trabajo.listar_hojas_trabajo'))
    
    presupuesto = get_by_id(Presupuesto, hoja.id_presupuesto)
    proyecto = get_by_id(Proyecto, presupuesto.id_proyecto)
    cliente = get_by_id(Cliente, proyecto.id_cliente) if proyecto.id_cliente else None
    
    # Obtener lista de proveedores para los selectores
    proveedores = get_all(Proveedor)
    
    # Obtener todos los proveedores asignados a partidas (versión temporal sin campos unitario/cantidad)
    proveedores_por_partida = {}
    
    # Aplicar el filtro de sanitizado al actualizar la hoja de trabajo
    if request.method == 'POST':
        try:
            data = {
                'tipo_via': request.form.get('tipo_via'),
                'nombre_via': request.form.get('nombre_via'),
                'numero_via': request.form.get('numero_via'),
                'puerta': request.form.get('puerta'),
                'codigo_postal': request.form.get('codigo_postal'),
                'poblacion': request.form.get('poblacion'),
                'titulo': request.form.get('titulo'),
                'notas': sanitize_html(request.form.get('notas', '')),
                'tecnico_encargado': request.form.get('tecnico_encargado') or 'Sin asignar',
                'estado': request.form.get('estado')
            }
            
            if request.form.get('aprobacion'):
                data['aprobacion'] = request.form.get('aprobacion')
                data['fecha_aprobacion'] = datetime.utcnow()
            
            update(HojaTrabajo, id, data)
            flash('Hoja de trabajo actualizada correctamente', 'success')
            return redirect(url_for('hojas_trabajo.editar_hoja', id=id))
        
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar hoja de trabajo: {str(e)}', 'danger')
    
    capitulos = CapituloHoja.query.filter_by(id_hoja=id).order_by(CapituloHoja.numero).all()
    partidas = PartidaHoja.query.filter_by(id_hoja=id).all()
    
    # Organizar partidas por capítulo
    partidas_por_capitulo = {}
    for partida in partidas:
        if partida.capitulo_numero not in partidas_por_capitulo:
            partidas_por_capitulo[partida.capitulo_numero] = []
        partidas_por_capitulo[partida.capitulo_numero].append(partida)
    
    # Para compatibilidad con la plantilla, aseguramos que todos los datos estén correctamente formateados
    for capitulo in capitulos:
        capitulo.partidas = partidas_por_capitulo.get(capitulo.numero, [])
    
    # Asegurar que la fecha está disponible en el formato correcto
    if hoja.fecha and not isinstance(hoja.fecha, str):
        hoja.fecha_str = hoja.fecha.strftime('%Y-%m-%d')
    
    # Si el técnico encargado no está definido, usar un valor por defecto
    if not hoja.tecnico_encargado:
        hoja.tecnico_encargado = request.form.get('tecnico_encargado', '') or ''
    
    return render_template('hojas_trabajo/editar_hoja_multiprov.html', 
                          hoja=hoja,
                          presupuesto=presupuesto,
                          proyecto=proyecto,
                          cliente=cliente,
                          capitulos=capitulos,
                          partidas_por_capitulo=partidas_por_capitulo,
                          proveedores=proveedores,
                          proveedores_por_partida=proveedores_por_partida)

@hojas_trabajo_bp.route('/eliminar/<int:id>', methods=['POST'])
def eliminar_hoja(id):
    """Elimina una hoja de trabajo"""
    try:
        hoja = get_by_id(HojaTrabajo, id)
        if not hoja:
            flash('Hoja de trabajo no encontrada', 'danger')
            return redirect(url_for('hojas_trabajo.listar_hojas_trabajo'))
            
        id_presupuesto = hoja.id_presupuesto
        
        delete(HojaTrabajo, id)
        flash('Hoja de trabajo eliminada correctamente', 'success')
        return redirect(url_for('hojas_trabajo.hojas_por_presupuesto', id_presupuesto=id_presupuesto))
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar hoja de trabajo: {str(e)}', 'danger')
        return redirect(url_for('hojas_trabajo.listar_hojas_trabajo'))

@hojas_trabajo_bp.route('/capitulo/nuevo/<int:id_hoja>', methods=['POST'])
def nuevo_capitulo(id_hoja):
    """Agrega un nuevo capítulo a la hoja de trabajo"""
    try:
        capitulo = CapituloHoja(
            id_hoja=id_hoja,
            numero=request.form.get('numero'),
            descripcion=request.form.get('descripcion')
        )
        db.session.add(capitulo)
        db.session.commit()
        flash('Capítulo creado correctamente', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al crear capítulo: {str(e)}', 'danger')
    
    return redirect(url_for('hojas_trabajo.editar_hoja', id=id_hoja))

@hojas_trabajo_bp.route('/partida/nueva/<int:id_hoja>', methods=['POST'])
def nueva_partida(id_hoja):
    """Agrega una nueva partida a la hoja de trabajo"""
    try:
        cantidad = float(request.form.get('cantidad') or 0)
        precio = float(request.form.get('precio') or 0)
        total = cantidad * precio
        margen = float(request.form.get('margen') or 0)
        final = total * (1 + margen / 100)
        
        # Datos del proveedor
        id_proveedor = request.form.get('id_proveedor')
        precio_proveedor = float(request.form.get('precio_proveedor') or 0)
        
        # Usar clean_html para preservar etiquetas HTML básicas
        descripcion = clean_html(request.form.get('descripcion', '').strip())
        
        partida = PartidaHoja(
            id_hoja=id_hoja,
            capitulo_numero=request.form.get('capitulo_numero'),
            descripcion=descripcion,
            unitario=request.form.get('unitario'),
            cantidad=cantidad,
            precio=precio,
            total=total,
            margen=margen,
            final=final,
            id_proveedor_principal=id_proveedor if id_proveedor else None,
            precio_proveedor=precio_proveedor if precio_proveedor > 0 else None
        )
        db.session.add(partida)
        db.session.flush()  # Para obtener el ID
        
        # Si hay un proveedor asignado, también se crea el registro en ProveedorPartida
        if id_proveedor and precio_proveedor > 0:
            proveedor_partida = ProveedorPartida(
                id_partida=partida.id,
                id_proveedor=id_proveedor,
                precio=precio_proveedor,
                margen_proveedor=0,  # Valor por defecto
                fecha_asignacion=datetime.utcnow(),
                estado='Pendiente'
            )
            # Calcular el final del proveedor
            proveedor_partida.calcular_final_proveedor()
            db.session.add(proveedor_partida)
        
        db.session.commit()
        flash('Partida creada correctamente', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al crear partida: {str(e)}', 'danger')
    
    return redirect(url_for('hojas_trabajo.editar_hoja', id=id_hoja))

@hojas_trabajo_bp.route('/pdf/<int:id>')
def generar_pdf(id):
    """Genera un PDF para la hoja de trabajo"""
    hoja = get_by_id(HojaTrabajo, id)
    if not hoja:
        flash('Hoja de trabajo no encontrada', 'danger')
        return redirect(url_for('hojas_trabajo.listar_hojas_trabajo'))
    
    presupuesto = get_by_id(Presupuesto, hoja.id_presupuesto)
    proyecto = get_by_id(Proyecto, presupuesto.id_proyecto)
    cliente = get_by_id(Cliente, proyecto.id_cliente) if proyecto.id_cliente else None
    
    capitulos = CapituloHoja.query.filter_by(id_hoja=id).order_by(CapituloHoja.numero).all()
    partidas = PartidaHoja.query.filter_by(id_hoja=id).all()
    
    # Organizar partidas por capítulo
    partidas_por_capitulo = {}
    for partida in partidas:
        if partida.capitulo_numero not in partidas_por_capitulo:
            partidas_por_capitulo[partida.capitulo_numero] = []
        partidas_por_capitulo[partida.capitulo_numero].append(partida)
    
    # Generar PDF
    pdf_file = generar_pdf_hoja_trabajo(hoja, proyecto, cliente, capitulos, partidas_por_capitulo)
    
    # Retornar el archivo PDF
    return send_file(
        pdf_file,
        as_attachment=True,
        download_name=f"HojaTrabajo_{hoja.referencia}.pdf",
        mimetype='application/pdf'
    )

@hojas_trabajo_bp.route('/ver-pdf/<int:id>')
def ver_pdf(id):
    """Genera y abre el PDF en el navegador sin descargarlo"""
    hoja = get_by_id(HojaTrabajo, id)
    if not hoja:
        flash('Hoja de trabajo no encontrada', 'danger')
        return redirect(url_for('hojas_trabajo.listar_hojas_trabajo'))
    
    presupuesto = get_by_id(Presupuesto, hoja.id_presupuesto)
    proyecto = get_by_id(Proyecto, presupuesto.id_proyecto)
    cliente = get_by_id(Cliente, proyecto.id_cliente) if proyecto.id_cliente else None
    
    capitulos = CapituloHoja.query.filter_by(id_hoja=id).order_by(CapituloHoja.numero).all()
    partidas = PartidaHoja.query.filter_by(id_hoja=id).all()
    
    # Organizar partidas por capítulo
    partidas_por_capitulo = {}
    for partida in partidas:
        if partida.capitulo_numero not in partidas_por_capitulo:
            partidas_por_capitulo[partida.capitulo_numero] = []
        partidas_por_capitulo[partida.capitulo_numero].append(partida)
    
    # Generar PDF
    pdf_file = generar_pdf_hoja_trabajo(hoja, proyecto, cliente, capitulos, partidas_por_capitulo)
    
    # Retornar el archivo PDF para visualización directa en el navegador
    return send_file(
        pdf_file,
        download_name=f"HojaTrabajo_{hoja.referencia}.pdf",
        mimetype='application/pdf',
        as_attachment=False
    )

@hojas_trabajo_bp.route('/exportar-excel/<int:id>')
def exportar_excel(id):
    """Exporta la hoja de trabajo a un archivo Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        flash('La librería openpyxl no está instalada. No se puede exportar a Excel.', 'danger')
        return redirect(url_for('hojas_trabajo.editar_hoja', id=id))
    
    hoja = get_by_id(HojaTrabajo, id)
    if not hoja:
        flash('Hoja de trabajo no encontrada', 'danger')
        return redirect(url_for('hojas_trabajo.listar_hojas_trabajo'))
    
    presupuesto = get_by_id(Presupuesto, hoja.id_presupuesto)
    proyecto = get_by_id(Proyecto, presupuesto.id_proyecto)
    cliente = get_by_id(Cliente, proyecto.id_cliente) if proyecto.id_cliente else None
    
    capitulos = CapituloHoja.query.filter_by(id_hoja=id).order_by(CapituloHoja.numero).all()
    partidas = PartidaHoja.query.filter_by(id_hoja=id).all()
    
    # Organizar partidas por capítulo
    partidas_por_capitulo = {}
    for partida in partidas:
        if partida.capitulo_numero not in partidas_por_capitulo:
            partidas_por_capitulo[partida.capitulo_numero] = []
        partidas_por_capitulo[partida.capitulo_numero].append(partida)
    
    # Crear libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja de Trabajo"
    
    # Estilos
    header_font = Font(name='Arial', size=12, bold=True)
    header_fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Información de cabecera
    ws['A1'] = "HOJA DE TRABAJO"
    ws['A1'].font = Font(name='Arial', size=16, bold=True)
    ws.merge_cells('A1:G1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws['A3'] = "Referencia:"
    ws['B3'] = hoja.referencia
    ws['A4'] = "Fecha:"
    ws['B4'] = hoja.fecha.strftime('%d/%m/%Y')
    ws['A5'] = "Presupuesto:"
    ws['B5'] = presupuesto.referencia
    ws['A6'] = "Proyecto:"
    ws['B6'] = proyecto.nombre_proyecto or proyecto.referencia
    ws['A7'] = "Cliente:"
    ws['B7'] = f"{cliente.nombre} {cliente.apellidos}" if cliente else "Sin cliente"
    ws['A8'] = "Técnico:"
    ws['B8'] = hoja.tecnico_encargado or "Sin asignar"
    
    # Agregar capítulos y partidas
    row = 10
    total_hoja = 0
    
    for capitulo in capitulos:
        # Título del capítulo
        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'] = f"CAPÍTULO {capitulo.numero}: {capitulo.descripcion}"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws[f'A{row}'].alignment = Alignment(horizontal='left')
        row += 1
        
        # Cabecera de partidas
        ws['A'+str(row)] = "Descripción"
        ws['E'+str(row)] = "Cantidad"
        ws['F'+str(row)] = "Precio (€)"
        ws['G'+str(row)] = "Total (€)"
        for col in ['A', 'E', 'F', 'G']:
            ws[col+str(row)].font = header_font
            ws[col+str(row)].border = border
        row += 1
        
        # Partidas del capítulo
        subtotal_capitulo = 0
        partidas_capitulo = partidas_por_capitulo.get(capitulo.numero, [])
        
        for partida in partidas_capitulo:
            ws.merge_cells(f'A{row}:D{row}')
            ws[f'A{row}'] = sanitize_html(partida.descripcion) if partida.descripcion else ''
            ws[f'E{row}'] = partida.cantidad
            ws[f'F{row}'] = partida.precio
            ws[f'G{row}'] = partida.final
            
            for col in ['A', 'E', 'F', 'G']:
                ws[col+str(row)].border = border
            
            subtotal_capitulo += partida.final or 0
            row += 1
        
        # Subtotal del capítulo
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "Subtotal Capítulo:"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].alignment = Alignment(horizontal='right')
        ws[f'G{row}'] = subtotal_capitulo
        ws[f'G{row}'].font = Font(bold=True)
        ws[f'G{row}'].border = border
        
        total_hoja += subtotal_capitulo
        row += 2  # Espacio entre capítulos
    
    # Total de la hoja
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "TOTAL HOJA DE TRABAJO:"
    ws[f'A{row}'].font = Font(name='Arial', size=12, bold=True)
    ws[f'A{row}'].alignment = Alignment(horizontal='right')
    ws[f'G{row}'] = total_hoja
    ws[f'G{row}'].font = Font(name='Arial', size=12, bold=True)
    ws[f'G{row}'].border = border
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    
    # Guardar en archivo temporal
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()
    
    # Enviar archivo
    return_data = send_file(
        temp_file.name,
        as_attachment=True,
        download_name=f"HojaTrabajo_{hoja.referencia}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Borrar archivo temporal después de enviar (registramos la función de limpieza)
    @return_data.call_on_close
    def delete_temp_file():
        try:
            os.unlink(temp_file.name)
        except:
            pass
    
    return return_data
